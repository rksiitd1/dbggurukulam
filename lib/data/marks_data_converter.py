import re
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import os

def parse_typescript_data(file_path):
    """Parse the TypeScript file and extract the data structure"""
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Try multiple patterns to find the marksData array
    patterns = [
        r'export const marksData: ClassData\[\] = (\[[\s\S]*?\]);',
        r'export const marksData:\s*ClassData\[\]\s*=\s*(\[[\s\S]*?\]);',
        r'const marksData:\s*ClassData\[\]\s*=\s*(\[[\s\S]*?\]);',
        r'marksData:\s*ClassData\[\]\s*=\s*(\[[\s\S]*?\]);'
    ]
    
    data_str = None
    for pattern in patterns:
        match = re.search(pattern, content)
        if match:
            data_str = match.group(1)
            break
    
    if not data_str:
        print("Could not find marksData with regex, trying manual parsing...")
        return manual_parse_data(content)
    
    # Clean up the TypeScript syntax to make it JSON-like
    # Remove comments first
    data_str = re.sub(r'//.*?\n', '\n', data_str)
    
    # Add quotes to unquoted keys
    data_str = re.sub(r'(\w+):\s*{', r'"\1": {', data_str)
    data_str = re.sub(r'(\w+):\s*(\d+|null)', r'"\1": \2', data_str)
    data_str = re.sub(r'(\w+):\s*"', r'"\1": "', data_str)
    data_str = re.sub(r'(\w+):\s*\[', r'"\1": [', data_str)
    
    try:
        data = json.loads(data_str)
        return data
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        print("Falling back to manual parsing...")
        # Manual parsing as fallback
        return manual_parse_data(content)

def manual_parse_data(content):
    """Manually parse the TypeScript data as a fallback"""
    # This is a more robust parser for the specific structure
    classes_data = []
    
    # Find the marksData array start
    start_pattern = r'export const marksData.*?=\s*\['
    start_match = re.search(start_pattern, content, re.DOTALL)
    
    if not start_match:
        # Try alternative patterns
        alt_patterns = [
            r'const marksData.*?=\s*\[',
            r'marksData.*?=\s*\['
        ]
        for pattern in alt_patterns:
            start_match = re.search(pattern, content, re.DOTALL)
            if start_match:
                break
    
    if not start_match:
        raise ValueError("Could not find marksData array in the file")
    
    # Find the matching closing bracket
    start_pos = start_match.end() - 1  # Position of the opening bracket
    bracket_count = 0
    end_pos = start_pos
    
    for i, char in enumerate(content[start_pos:], start_pos):
        if char == '[' or char == '{':
            bracket_count += 1
        elif char == ']' or char == '}':
            bracket_count -= 1
            if bracket_count == 0 and char == ']':
                end_pos = i
                break
    
    # Extract the array content
    array_content = content[start_pos + 1:end_pos]
    
    # Split by class objects (look for className pattern)
    class_sections = re.split(r'(?=\s*{\s*className)', array_content)
    
    for section in class_sections:
        if 'className' not in section:
            continue
            
        # Extract class name
        class_match = re.search(r'className:\s*["\']([^"\']+)["\']', section)
        if not class_match:
            continue
        class_name = class_match.group(1)
        
        # Extract students array
        students_match = re.search(r'students:\s*\[(.*)\]', section, re.DOTALL)
        if not students_match:
            continue
        students_content = students_match.group(1)
        
        students = []
        
        # Split students by objects
        student_sections = re.split(r'(?=\s*{\s*rollNo)', students_content)
        
        for student_section in student_sections:
            if 'rollNo' not in student_section:
                continue
                
            # Extract roll number
            roll_match = re.search(r'rollNo:\s*(\d+)', student_section)
            if not roll_match:
                continue
            roll_no = int(roll_match.group(1))
            
            # Extract name
            name_match = re.search(r'name:\s*["\']([^"\']+)["\']', student_section)
            if not name_match:
                continue
            name = name_match.group(1)
            
            # Extract subjects
            subjects_match = re.search(r'subjects:\s*{(.*)}', student_section, re.DOTALL)
            if not subjects_match:
                continue
            subjects_content = subjects_match.group(1)
            
            subjects = {}
            
            # Find all subject entries
            subject_pattern = r'(\w+):\s*{\s*written:\s*(null|\d+),\s*oral:\s*(null|\d+),\s*project:\s*(null|\d+)\s*}'
            subject_matches = re.findall(subject_pattern, subjects_content)
            
            for subject_name, written, oral, project in subject_matches:
                subjects[subject_name] = {
                    'written': None if written == 'null' else int(written),
                    'oral': None if oral == 'null' else int(oral),
                    'project': None if project == 'null' else int(project)
                }
            
            if subjects:  # Only add student if they have subjects
                students.append({
                    'rollNo': roll_no,
                    'name': name,
                    'subjects': subjects
                })
        
        if students:  # Only add class if it has students
            classes_data.append({
                'className': class_name,
                'students': students
            })
    
    return classes_data

def create_excel_file(data, output_path):
    """Create a comprehensive Excel file with multiple sheets"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_data = []
    
    for class_data in data:
        class_name = class_data['className']
        students = class_data['students']
        
        # Calculate statistics
        total_students = len(students)
        students_with_marks = len([s for s in students if any(
            any(mark is not None for mark in subject.values()) 
            for subject in s['subjects'].values()
        )])
        
        # Get subjects for this class
        subjects = set()
        for student in students:
            subjects.update(student['subjects'].keys())
        subjects = sorted(subjects)
        
        summary_data.append([
            class_name,
            total_students,
            students_with_marks,
            ', '.join(subjects)
        ])
    
    # Add summary headers
    summary_headers = ['Class', 'Total Students', 'Students with Marks', 'Subjects']
    summary_ws.append(summary_headers)
    
    # Style summary headers
    for col in range(1, len(summary_headers) + 1):
        cell = summary_ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Add summary data
    for row_data in summary_data:
        summary_ws.append(row_data)
        row_num = summary_ws.max_row
        for col in range(1, len(row_data) + 1):
            cell = summary_ws.cell(row_num, col)
            cell.border = border
            cell.alignment = center_alignment
    
    # Auto-adjust column widths for summary
    for col in summary_ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        summary_ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Create individual class sheets
    for class_data in data:
        class_name = class_data['className']
        students = class_data['students']
        
        # Create worksheet
        ws = wb.create_sheet(f"Class {class_name}")
        
        # Get all subjects for this class
        all_subjects = set()
        for student in students:
            all_subjects.update(student['subjects'].keys())
        all_subjects = sorted(all_subjects)
        
        # Create headers
        headers = ['Roll No', 'Name']
        for subject in all_subjects:
            headers.extend([f'{subject} (W)', f'{subject} (O)', f'{subject} (P)', f'{subject} Total'])
        headers.append('Grand Total')
        
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Add student data
        for student in students:
            row_data = [student['rollNo'], student['name']]
            grand_total = 0
            
            for subject in all_subjects:
                if subject in student['subjects']:
                    marks = student['subjects'][subject]
                    written = marks['written'] or 0
                    oral = marks['oral'] or 0
                    project = marks['project'] or 0
                    subject_total = written + oral + project
                    
                    row_data.extend([
                        written if marks['written'] is not None else '',
                        oral if marks['oral'] is not None else '',
                        project if marks['project'] is not None else '',
                        subject_total if any(m is not None for m in marks.values()) else ''
                    ])
                    
                    if any(m is not None for m in marks.values()):
                        grand_total += subject_total
                else:
                    row_data.extend(['', '', '', ''])
            
            row_data.append(grand_total if grand_total > 0 else '')
            ws.append(row_data)
            
            # Style data row
            row_num = ws.max_row
            for col in range(1, len(row_data) + 1):
                cell = ws.cell(row_num, col)
                cell.border = border
                cell.alignment = center_alignment
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 15)
    
    # Save Excel file
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")

def create_pdf_file(data, output_path):
    """Create a comprehensive PDF report"""
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    elements.append(Paragraph("Student Marks Report", title_style))
    elements.append(Spacer(1, 20))
    
    # Summary section
    elements.append(Paragraph("Summary", styles['Heading2']))
    summary_data = [['Class', 'Total Students', 'Students with Marks', 'Subjects']]
    
    for class_data in data:
        class_name = class_data['className']
        students = class_data['students']
        
        total_students = len(students)
        students_with_marks = len([s for s in students if any(
            any(mark is not None for mark in subject.values()) 
            for subject in s['subjects'].values()
        )])
        
        subjects = set()
        for student in students:
            subjects.update(student['subjects'].keys())
        subjects_str = ', '.join(sorted(subjects))
        
        summary_data.append([class_name, str(total_students), str(students_with_marks), subjects_str])
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(PageBreak())
    
    # Detailed class reports
    for class_data in data:
        class_name = class_data['className']
        students = class_data['students']
        
        # Class title
        elements.append(Paragraph(f"Class {class_name} - Detailed Marks", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        # Get all subjects for this class
        all_subjects = set()
        for student in students:
            all_subjects.update(student['subjects'].keys())
        all_subjects = sorted(all_subjects)
        
        # Create table data
        table_data = []
        
        # Headers
        headers = ['Roll', 'Name']
        for subject in all_subjects:
            headers.extend([f'{subject}(W)', f'{subject}(O)', f'{subject}(P)', f'{subject} Tot'])
        headers.append('Total')
        table_data.append(headers)
        
        # Student data
        for student in students:
            row = [str(student['rollNo']), student['name']]
            grand_total = 0
            
            for subject in all_subjects:
                if subject in student['subjects']:
                    marks = student['subjects'][subject]
                    written = marks['written'] if marks['written'] is not None else '-'
                    oral = marks['oral'] if marks['oral'] is not None else '-'
                    project = marks['project'] if marks['project'] is not None else '-'
                    
                    # Calculate subject total
                    subject_total = 0
                    if marks['written'] is not None:
                        subject_total += marks['written']
                    if marks['oral'] is not None:
                        subject_total += marks['oral']
                    if marks['project'] is not None:
                        subject_total += marks['project']
                    
                    row.extend([str(written), str(oral), str(project), str(subject_total) if subject_total > 0 else '-'])
                    
                    if subject_total > 0:
                        grand_total += subject_total
                else:
                    row.extend(['-', '-', '-', '-'])
            
            row.append(str(grand_total) if grand_total > 0 else '-')
            table_data.append(row)
        
        # Create and style table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Add page break between classes (except for the last one)
        if class_data != data[-1]:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements)
    print(f"PDF file saved: {output_path}")

def main():
    """Main function to process the TypeScript file"""
    input_file = "marks-data.ts"
    excel_output = "student_marks.xlsx"
    pdf_output = "student_marks_report.pdf"
    
    try:
        # Check if input file exists
        if not os.path.exists(input_file):
            print(f"Error: {input_file} not found in the current directory")
            return
        
        # Parse the TypeScript data
        print("Parsing TypeScript data...")
        data = parse_typescript_data(input_file)
        
        # Create Excel file
        print("Creating Excel file...")
        create_excel_file(data, excel_output)
        
        # Create PDF file
        print("Creating PDF file...")
        create_pdf_file(data, pdf_output)
        
        print("\nConversion completed successfully!")
        print(f"Files created:")
        print(f"- {excel_output}")
        print(f"- {pdf_output}")
        
        # Print some statistics
        total_students = sum(len(class_data['students']) for class_data in data)
        total_classes = len(data)
        print(f"\nStatistics:")
        print(f"- Total classes: {total_classes}")
        print(f"- Total students: {total_students}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
